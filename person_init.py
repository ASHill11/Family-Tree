import openpyxl
import dbscripts

class Person:
    def __init__(self, first_name, last_name, class_year, parse_id, parents=[], children=[]):
        self.parse_id = parse_id
        self.first_name = first_name
        self.last_name = last_name
        self.class_year = class_year
        self.parents = parents
        self.children = children


def read_excel_data(filename):
    """
    Reads data from an Excel file and returns a list of Person objects
    """
    # Load the workbook and select the first worksheet
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    # Initialize a list to hold the Person objects
    people_list = []

    # Loop through each row in the worksheet until an empty row is reached
    row_num = 2  # start at row 2 to skip the header row
    while worksheet.cell(row=row_num, column=1).value is not None:
        # Get the data from the row and create a Person object
        id_num = worksheet.cell(row=row_num, column=1).value
        first_name = worksheet.cell(row=row_num, column=2).value
        last_name = worksheet.cell(row=row_num, column=3).value
        class_year = worksheet.cell(row=row_num, column=4).value

        parents = []
        for i in [5, 6, 7]:
            if worksheet.cell(row=row_num, column=i).value:
                parents.append(worksheet.cell(row=row_num, column=i).value)

        children = []
        for i in [8, 9, 10, 11, 12]:
            if worksheet.cell(row=row_num, column=i).value:
                children.append(worksheet.cell(row=row_num, column=i).value)
        person = Person(id_num, first_name, last_name, class_year, parents, children)

        # Add the Person object to the list
        people_list.append(person)

        # Move to the next row
        row_num += 1

    # Close the workbook and return the list of Person objects
    workbook.close()
    return people_list


def create_name_dict(people):
    name_dict = {}
    for person in people:
        class_year = ''
        if person.class_year is not None:
            class_year = str(int(str(person.class_year).replace('?', '')) % 100)
            class_year = ' \'' + '0' * (2 - len(class_year)) + class_year
        full_name = person.first_name + ' ' + person.last_name + class_year
        name_dict[person.parse_id] = full_name
    return name_dict


def create_person_id_dict(people):
    id_person_dict = {}
    for person in people:
        id_person_dict[person.parse_id] = person
    return id_person_dict

def initialize_people():
    dbconn = dbscripts.create_connection()
    # people = dbscripts.get_people_from_db(dbconn)
    people = [] # comment this out and uncomment the previous line to only read the db on startup
    if len(people) == 0:
        print("Populating database from excel file.")
        people = read_excel_data('family-tree-data.xlsx')
        dbscripts.add_people_to_db(dbconn, people)
    person_name_dict = create_name_dict(people)
    person_id_dict = create_person_id_dict(people)
    if dbconn is not None: dbconn.close()
    return (people, person_name_dict, person_id_dict)